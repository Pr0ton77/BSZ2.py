import tkinter
import openpyxl
import time

file_name = "Name"

wb = openpyxl.load_workbook(f"BDs/{file_name}.xlsx")
ws = wb.active

def clear():
    text_info.delete("1.0", "end")
    wind.update()
    time.sleep(0.1)

def speed(i,line, info):
    Ninfo = ws[f"{line}{i}"].value
    if (line == "Q") or (line == "P") or (line == "O") or (line == "N"):
        if ws[f"{line}{i}"].value == None:
            Ninfo = "Ученик не выбывал/прибывал"

    elif line == "H":
        if (ws[f"{line}{i}"].value == " ") or (ws[f"{line}{i}"].value == None):
            Ninfo = "Пусто"
    elif line == "I":
        if (ws[f"{line}{i}"].value == " ") or (ws[f"{line}{i}"].value == None):
            Ninfo = "-"

    text_info.insert('1.0', f'\n┌───────────────────────────────────────────────\n{info}:\n{Ninfo}\n└───────────────────────────────────────────────\n\n')
    time.sleep(0.1)
    wind.update()


def Nm(None_):
    i = 1
    ws22 = ""
    while True:
        if file_name == "Name":
            text_info.configure(state = tkinter.NORMAL)
            clear()
            text_info.insert("1.0","Ошибка!\n\nВведите название файлка БВ\nво вторую строку")
            text_info.configure(state = tkinter.DISABLED)
            break
        ww = ws[f"C{i}"].value

        if ws[f"C{i}"].value != None:
            ww = ww.title().strip()
        try:
            ws2 = ww.split()
            del(ws2[-1])
            ws22 = f"{ws2[0]} {ws2[1]}"
        except:
            pass

        if (Ented.get().title().strip() == ww) or (Ented.get().title().strip() == ws22.title().strip()):
            text_info.configure(state=tkinter.NORMAL)
            clear()
            speed(i,"U","Родной язык")
            speed(i,"T","Свидетельство о рождении")
            speed(i,"S","Страховой полис")
            speed(i,"R","Снилс")

            speed(i,"Q","В какую школу выбыл")
            speed(i,"P","Когда выбыл")
            speed(i,"O","Из какой школы прибыл")
            speed(i,"N","Когда прибыл")

            speed(i,"M","Телефон родителей")
            speed(i,"L","Адрес фактического проживания")
            speed(i,"K","Место работы родителей и должность")
            speed(i,"J","Ф.И.О родителей")
            speed(i,"I","Соц.Статус ( неполная/многодетная семья/кол-во детей)")

            speed(i,"H","Инв. ОВЗ")
            speed(i,"G","Национальность")
            speed(i,"F","Число и мясяц рождения")
            speed(i,"E","Пол")
            speed(i,"D","Класс")

            text_info.configure(state=tkinter.DISABLED)
            break
        else:
            i += 1
        if ws[f"C{i}"].value == None:
            if ws[f"C{i + 1}"].value == None:
                if ws[f"C{i + 2}"].value == None:
                    if ws[f"C{i + 3}"].value == None:
                        if ws[f"C{i + 4}"].value == None:
                            text_info.configure(state=tkinter.NORMAL)
                            clear()
                            text_info.insert("1.0", f"\nПользователь \n\"{Ented.get().title()}\" \nне найден\n\nПопробуйте снова")
                            text_info.configure(state=tkinter.DISABLED)
                            break

def All_list(None_):
    i2 = 2
    sh = 1

    text_info.configure(state=tkinter.NORMAL)
    clear()
    Nn = None
    while True:
        if file_name == "Name":
            text_info.configure(state = tkinter.NORMAL)
            clear()
            text_info.insert("1.0","Ошибка!\n\nВведите название файлка БВ\nво вторую строку")
            text_info.configure(state = tkinter.DISABLED)
            break
        if ws[f"C{i2}"].value == Nn:
            if ws[f"C{i2+1}"].value == Nn:
                if ws[f"C{i2+2}"].value == Nn:
                    if ws[f"C{i2+3}"].value == Nn:
                        if ws[f"C{i2+4}"].value == Nn:
                            break

        text_info.insert("1.0",f'{sh}: {ws[f"C{i2}"].value} {ws[f"D{i2}"].value}\n')
        sh += 1
        wind.update()
        i2 += 1
        time.sleep(0)
    text_info.configure(state=tkinter.DISABLED)

def file_writed(None_):
    global file_name
    global ws
    global wb
    try:
        file_name = file_write.get()
        wb = openpyxl.load_workbook(f"BDs/{file_name}.xlsx")
        ws = wb.active
        text_info.configure(state=tkinter.NORMAL)
        clear()
        text_info.insert("1.0","Отлично!\n\nВаш файл был успешно записан")

    except:
        text_info.configure(state=tkinter.NORMAL)
        clear()
        text_info.insert("1.0","Ошибка! \n\nНеправильно введено\nназвание файла\nлибо его вовсе нет")
        text_info.configure(state=tkinter.DISABLED)

wind = tkinter.Tk()
wind.title("LIST")
wind.geometry("400x600")

text_info = tkinter.Text(width=48,height=30,state=tkinter.DISABLED)
text_info.place(x=1,y=100)

scroll = tkinter.Scrollbar(command= text_info.yview)
scroll.place(y=100,x=385, height=485)
text_info.config( yscrollcommand= scroll.set)
wind.update()

but = tkinter.Button(text="Ввод",width=15,height=2,font = ", 6", bg="#AD2222")
but.bind("<Button-1>", Nm)
but.place(x=300,y=20)

but_2 = tkinter.Button(text="Весь список уч.",width=15,height=2,font = ", 6", bg="#AD2222")
but_2.bind("<Button-1>", All_list)
but_2.place(x=300,y=52)

Ented = tkinter.Entry(font = ", 10")
Ented.place(x=1,y=20,width=284,height=30)

file_write = tkinter.Entry()
file_write.place(x=1,y=55,width=70,height=25)

but_3 = tkinter.Button(text="Ввод", bg="#AD2222")
but_3.bind("<Button-1>",file_writed)
but_3.place(x=75,y=55)

wind.resizable(False, False)
wind.configure(bg='black')
wind.mainloop()